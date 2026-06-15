"""Build dijital-ikiz-secim.xlsx — checklist of all visualizable survey questions."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path

OUT = Path(r"C:\Users\furkan\Documents\dev\dairymind\dijital-ikiz-secim.xlsx")

# Data: (Soru No, Bölüm, Soru, Kategori, Görsel/Animasyon, Karmaşıklık, Öncelik 1-5)
ROWS = [
    # --- Section 1 ---
    ("1.7", "1·İşletme", "GPS koordinatları", "Konum",
     "Harita pin işareti (köşede mini Türkiye haritası, pin'in koordinata göre konumu)", "Orta", 2),

    # --- Section 2: ARAZİ & ALTYAPI (visual gold mine) ---
    ("2.1", "2·Arazi", "Toplam arazi büyüklüğü", "Arazi",
     "Yeşil çim alan boyutu ölçeklenir (küçük/orta/büyük)", "Kolay", 4),
    ("2.4", "2·Arazi", "Tarım/yem arazisi (dönüm)", "Arazi",
     "Çiftliğin yanına yeşil tarla parseli (mısır/yonca dokulu)", "Orta", 4),
    ("2.5", "2·Arazi", "Sulama imkanı", "Tarla",
     "Tarla üzerinde sprinkler/su fışkırma animasyonu", "Orta", 3),
    ("2.6", "2·Arazi", "Su kaynağı tipi (çoklu)", "Yapı",
     "Her tip için ayrı: kuyu/gölet/şebeke/dere/yağmur toplama tankı (multi gösterim)", "Zor", 5),
    ("2.7", "2·Arazi", "Ahır tipi", "Ana Yapı",
     "Serbest/bağlı/açık padok için 3 farklı ahır modeli", "Orta", 5),
    ("2.8", "2·Arazi", "Ahır toplam alanı (m²)", "Ölçek",
     "Ahır SVG'sinin boyutu m²'ye göre ölçeklenir", "Orta", 3),
    ("2.9", "2·Arazi", "Ahır kapasitesi (baş)", "Ölçek",
     "Görsel uzunluk + iç bölme sayısı kapasiteye göre", "Orta", 3),
    ("2.10", "2·Arazi", "Ahır yapım yılı", "Detay",
     "Eski (50+ yıl) ise weathered/eskimiş tonlar, yeni ise parlak", "Zor", 2),
    ("2.12", "2·Arazi", "Soğutma sistemi", "Donanım",
     "Ahır üstünde duş başlıkları + su zerresi animasyonu", "Orta", 4),
    ("2.13", "2·Arazi", "Aydınlatma tipi", "Detay",
     "LED parlak beyaz, halojen sarımsı, doğal hiç — pencere parıltısı farklı", "Kolay", 3),
    ("2.14", "2·Arazi", "Günlük aydınlatma süresi (saat)", "Atmosfer",
     "Gündüz/gece döngüsü — saat sayısına göre güneş pozisyonu", "Zor", 2),
    ("2.15", "2·Arazi", "Havalandırma tipi", "Donanım",
     "Doğal=büyük pencereler, mekanik=büyük fanlar (4-6 adet dönen)", "Orta", 4),
    ("2.16", "2·Arazi", "Yataklık materyali (çoklu)", "Detay",
     "Ahır içinde renkli zemin: sap=sarı, kum=bej, lastik mat=siyah", "Orta", 3),
    ("2.17", "2·Arazi", "Sağım odası tipi", "Ana Yapı",
     "Balık sırtı/karusel(dönen)/robotik/sabit boru ayrı SVG'ler", "Orta", 5),
    ("2.18", "2·Arazi", "Sağım kapasitesi (duraklı)", "Ölçek",
     "Sağım odası genişliği + iç durak sayısı görünür", "Orta", 3),
    ("2.19", "2·Arazi", "Süt tankı kapasitesi (L)", "Yapı",
     "Tank boyutu küçük/orta/büyük (3 scale)", "Kolay", 5),
    ("2.20", "2·Arazi", "Tank tipi (DX/IBT/pre-cooler)", "Detay",
     "Tank şekli ve borulama farklı (yatay/dikey, soğutma birimi yan)", "Orta", 3),
    ("2.21", "2·Arazi", "TMR mikser", "Araç",
     "Sarı yem karıştırma traktörü (dönen drum animasyonu)", "Orta", 4),
    ("2.22", "2·Arazi", "Mikser hacmi (m³)", "Ölçek",
     "Traktör boyutu hacme göre", "Kolay", 2),
    ("2.23", "2·Arazi", "Silaj alanı/silosu", "Yapı",
     "Beton silo / yığın silaj / balya — 3 farklı görsel", "Orta", 5),
    ("2.24", "2·Arazi", "Silaj kapasitesi (ton)", "Ölçek",
     "Silo yüksekliği veya yığın hacmi ton'a göre", "Kolay", 3),
    ("2.25", "2·Arazi", "Elektrik sistemi", "Altyapı",
     "Şebeke=elektrik direği, jeneratör=küçük yeşil kutu", "Orta", 4),
    ("2.26", "2·Arazi", "Yıllık elektrik tüketimi (kWh)", "Etiket",
     "Köşede dijital meter göstergesi (animasyonlu sayaç)", "Orta", 2),
    ("2.27", "2·Arazi", "Yenilenebilir enerji", "Altyapı",
     "Güneş paneli grup, biyogaz kupa, rüzgar türbini (3 görsel)", "Orta", 5),
    ("2.28", "2·Arazi", "Atık/gübre yönetimi", "Yapı",
     "Lagün/kompost(duman)/biyogaz tesisi (3 ayrı görsel)", "Orta", 5),

    # --- Section 3: SÜRÜ ---
    ("3.1", "3·Sürü", "Toplam baş sayısı", "Hayvan",
     "İnek figürleri (4'e 1 oranla), max 14-18 görünür", "Kolay", 5),
    ("3.2", "3·Sürü", "Sağmal inek sayısı", "Hayvan",
     "Ayrı grup — yetişkin inekler (büyük), beyaz-siyah", "Orta", 4),
    ("3.3", "3·Sürü", "Kuru dönem inek sayısı", "Hayvan",
     "Ayrı bölmede yatan ağır inekler", "Orta", 3),
    ("3.4", "3·Sürü", "Düve sayısı", "Hayvan",
     "Küçük boy inekler, genç görünümlü", "Orta", 4),
    ("3.5", "3·Sürü", "Buzağı sayısı", "Hayvan",
     "Bebek boy inekler, padokta ayrı bölme", "Orta", 4),
    ("3.6", "3·Sürü", "Boğa sayısı", "Hayvan",
     "Büyük boy boğa figürü (yaka rengi farklı)", "Kolay", 3),
    ("3.7", "3·Sürü", "Baskın ırk", "Detay",
     "Holstein=siyah-beyaz, Simental=kırmızı-beyaz, Jersey=kahve", "Orta", 4),
    ("3.8", "3·Sürü", "Irk dağılımı yüzdesi", "Detay",
     "İnek renk dağılımı yüzdeye göre karışım", "Zor", 3),
    ("3.18", "3·Sürü", "Grup ayrımı (çoklu)", "Yerleşim",
     "Ahır içinde renkli bölmeler (sağmal/kuru/düve/buzağı)", "Zor", 3),

    # --- Section 4: SAĞLIK & ÜREME ---
    ("4.1", "4·Sağlık", "Veteriner ziyaret sıklığı", "Araç",
     "Beyaz vet aracı park halinde (haftalık=her zaman görünür)", "Orta", 3),
    ("4.2", "4·Sağlık", "Aşı programı", "İkon",
     "İneklerin yanında küçük + tıbbi simge", "Kolay", 2),
    ("4.3", "4·Sağlık", "Hangi aşılar (çoklu)", "Etiket",
     "Köşede aşı listesi badge'leri", "Kolay", 2),
    ("4.7", "4·Sağlık", "Tırnak bakım sıklığı", "İkon",
     "Inek bacaklarında bakım yapan kişi figürü (mevsimsel)", "Zor", 1),
    ("4.10", "4·Sağlık", "Tohumlama tipi", "İkon",
     "SI=mavi simge, boğa=erkek figürü, ET=advanced badge", "Orta", 2),
    ("4.14", "4·Sağlık", "Östrus tespit (çoklu)", "Donanım",
     "İnek ayağında pedometre + ahırda kamera ışıkları", "Orta", 3),
    ("4.19", "4·Sağlık", "Kolostrum protokolü", "İkon",
     "Buzağı yanında biberon simgesi", "Kolay", 1),

    # --- Section 5: SÜT ---
    ("5.1", "5·Süt", "Yıllık toplam süt (ton)", "Etiket",
     "Köşede süt üretim göstergesi (ton/yıl + grafik)", "Kolay", 3),
    ("5.2", "5·Süt", "İnek başı günlük süt (kg)", "Animasyon",
     "Sağım odasından çıkan süt damla animasyonu — yoğunluk verime göre", "Orta", 4),
    ("5.4", "5·Süt", "Pik laktasyon günü", "Etiket",
     "Mini Wood eğrisi grafiği panel köşede", "Zor", 2),
    ("5.6", "5·Süt", "Süt yağ oranı (%)", "Etiket",
     "Tank üzerinde kalite badge (yağ %)", "Kolay", 2),
    ("5.7", "5·Süt", "Süt protein (%)", "Etiket",
     "Tank üzerinde protein % badge", "Kolay", 2),
    ("5.9", "5·Süt", "SHS (Somatik Hücre)", "Sağlık",
     "Yeşil/sarı/kırmızı sağlık göstergesi (200K altı yeşil)", "Orta", 4),
    ("5.10", "5·Süt", "Toplam bakteri sayısı", "Sağlık",
     "Hijyen rozeti (CFU bandı renkli)", "Orta", 2),
    ("5.13", "5·Süt", "Süt teslim sıklığı", "Araç",
     "Süt tankerinin gelme animasyonu (günlük/günaşırı)", "Orta", 4),
    ("5.16", "5·Süt", "Kalite primi", "Etiket",
     "Altın yıldız rozeti tank üstünde", "Kolay", 2),

    # --- Section 6: BESLENME ---
    ("6.1", "6·Beslenme", "Beslenme sistemi (TMR/komponent/mera)", "Donanım",
     "TMR=mikser görünür, mera=otlakta inekler", "Orta", 4),
    ("6.4", "6·Beslenme", "Mısır silajı üretimi (ton)", "Tarla",
     "Mısır tarlası SVG (ton'a göre alan büyür)", "Orta", 4),
    ("6.5", "6·Beslenme", "Yonca üretimi (ton)", "Tarla",
     "Yonca tarlası (mor çiçekli)", "Orta", 3),
    ("6.6", "6·Beslenme", "Kuru ot üretimi", "Tarla",
     "Buğday/saman tarlası (sarı) veya saman balyaları", "Orta", 3),
    ("6.10", "6·Beslenme", "Soya küspesi (ton)", "Depo",
     "Yem deposunda soya çuvalları", "Kolay", 2),
    ("6.13", "6·Beslenme", "Aylık yem maliyeti (₺)", "Etiket",
     "Yem deposu üstünde fiyat etiketi", "Kolay", 2),
    ("6.17", "6·Beslenme", "Su tüketimi takibi", "Donanım",
     "İnek bölmesi yanında su yalağı + sensör ışığı", "Orta", 2),
    ("6.18", "6·Beslenme", "İnek başı su tüketimi (L)", "Etiket",
     "Su sayacı L/gün göstergesi", "Kolay", 2),
    ("6.19", "6·Beslenme", "Yem deposu kapasitesi (ay)", "Yapı",
     "Ahır yanında yem deposu binası (büyüklük ay'a göre)", "Orta", 3),
    ("6.27", "6·Beslenme", "Otomatik yemleme", "Donanım",
     "Otomatik yem robotu (hareketli)", "Zor", 3),

    # --- Section 7: SAĞIM ---
    ("7.1", "7·Sağım", "Günlük sağım sayısı (2x/3x/robot)", "Animasyon",
     "Robot=sürekli yanıp sönen LED, 2x=günde 2 küme animasyon", "Orta", 3),
    ("7.4", "7·Sağım", "Vakum basıncı (kPa)", "Etiket",
     "Sağım odası üstünde dijital basınç göstergesi", "Kolay", 2),
    ("7.7", "7·Sağım", "Sağım öncesi temizlik (predip)", "İkon",
     "Sağım odası girişinde temizlik istasyonu (sprey ikonu)", "Orta", 2),
    ("7.8", "7·Sağım", "Sağım sonrası daldırma (postdip)", "İkon",
     "Sağım odası çıkışında ikinci temizlik istasyonu", "Orta", 2),
    ("7.12", "7·Sağım", "Auto-takeoff", "Detay",
     "Sağım ünitelerinde robotik kol simgesi", "Orta", 2),
    ("7.13", "7·Sağım", "Süt akış sayacı (her durakta)", "Donanım",
     "Her stall'da dijital ekran (akan rakam)", "Zor", 3),
    ("7.14", "7·Sağım", "İletkenlik ölçümü", "Sağlık",
     "Sağım ünitesinde mastitis erken uyarı LED", "Orta", 2),

    # --- Section 8: İŞGÜCÜ ---
    ("8.1", "8·İşgücü", "Tam zamanlı çalışan sayısı", "Kişi",
     "İşçi figürleri (1'e 1 oranla, max 6 görünür)", "Orta", 4),
    ("8.4", "8·İşgücü", "Sağımcı sayısı", "Kişi",
     "Sağım odası önünde sağımcı figürü (önlüklü)", "Orta", 3),
    ("8.5", "8·İşgücü", "Bakıcı/yemleme sayısı", "Kişi",
     "Yem deposu yanında işçi figürü", "Orta", 2),
    ("8.6", "8·İşgücü", "Tarla çalışanı sayısı", "Kişi",
     "Tarlada işçi figürü (traktör başında)", "Orta", 2),
    ("8.7", "8·İşgücü", "Veteriner durumu", "Kişi",
     "Beyaz önlüklü vet figürü (kadrolu=her zaman, dış=ara sıra)", "Orta", 3),
    ("8.8", "8·İşgücü", "Ziraat mühendisi", "Kişi",
     "Sarı baretli mühendis figürü", "Orta", 2),
    ("8.9", "8·İşgücü", "İşletme müdürü", "Kişi",
     "Tablet/clipboard tutan müdür figürü", "Orta", 2),

    # --- Section 9: FİNANS ---
    ("9.1", "9·Finans", "Yıllık ciro", "Etiket",
     "Köşede ₺ rozeti (ciro aralığına göre büyüklük)", "Kolay", 2),
    ("9.13", "9·Finans", "Devlet destekleri (çoklu)", "Etiket",
     "Çiftlik girişinde TKDK/TARSIM bayrağı/rozeti", "Kolay", 2),

    # --- Section 10: PAZARLAMA ---
    ("10.1", "10·Pazarlama", "Süt alıcısı tipi", "Araç",
     "Süt tankeri markalı (Pınar/Sütaş/birlik/yerel)", "Orta", 3),
    ("10.10", "10·Pazarlama", "Düve satışı", "Araç",
     "Hayvan satış kamyonu (düve ihracı)", "Orta", 2),
    ("10.11", "10·Pazarlama", "Gübre satışı", "Araç",
     "Gübre yükleyen kamyon (compost yanında)", "Orta", 2),
    ("10.13", "10·Pazarlama", "Markalı süt işleme", "Yapı",
     "Mini işleme tesisi binası (peynir/yoğurt simgeleri)", "Zor", 3),

    # --- Section 11: TEKNOLOJİ (sensor heaven) ---
    ("11.1", "11·Teknoloji", "Sürü yönetim yazılımı", "Etiket",
     "Köşede laptop/monitor ekranı (canlı veri akışı)", "Orta", 3),
    ("11.3", "11·Teknoloji", "Pedometre", "Donanım",
     "İneklerin ayağında küçük sarı bant + pulse LED", "Orta", 4),
    ("11.4", "11·Teknoloji", "Bolus/iç sıcaklık sensörü", "Donanım",
     "İnek üstünde küçük kırmızı sensör nokta (pulse)", "Orta", 3),
    ("11.5", "11·Teknoloji", "Otomatik tartım", "Donanım",
     "Yürüyüş yolunda tartım platformu (digital display)", "Orta", 2),
    ("11.7", "11·Teknoloji", "Hava istasyonu", "Donanım",
     "Çiftlik kenarında meteoroloji istasyonu (THI verisi)", "Orta", 3),
    ("11.8", "11·Teknoloji", "CO₂/NH₃ sensör", "Donanım",
     "Ahır içinde küçük sensör noktaları (yeşil/kırmızı LED)", "Orta", 3),
    ("11.9", "11·Teknoloji", "Akıllı su sayacı", "Donanım",
     "Su deposu üstünde dijital sayaç", "Kolay", 2),
    ("11.10", "11·Teknoloji", "Internet (fiber/4G)", "Altyapı",
     "Wifi/anten kulesi (fiber=parlak, 4G=zayıf)", "Kolay", 3),
    ("11.12", "11·Teknoloji", "Çiftlik kameraları", "Donanım",
     "Ahır köşelerinde CCTV kameralar (kırmızı LED)", "Orta", 3),
    ("11.13", "11·Teknoloji", "Mobil uygulama kullanımı", "Etiket",
     "Köşede telefon ikonu (mobil bağlantı göstergesi)", "Kolay", 1),
    ("11.15", "11·Teknoloji", "Buzağı doğum kamerası", "Donanım",
     "Buzağılama bölmesi üstünde kamera", "Kolay", 2),
    ("11.17", "11·Teknoloji", "Otomatik mineral dozlama", "Donanım",
     "Yem yolunda küçük dozaj robotları", "Zor", 2),

    # --- Section 12: SÜRDÜRÜLEBİLİRLİK ---
    ("12.1", "12·Çevre", "Karbon ayak izi (sertifikalı)", "Etiket",
     "Çiftlik üstünde 'Carbon Neutral' rozeti (yeşil yaprak)", "Kolay", 3),
    ("12.6", "12·Çevre", "GES gücü (kWp)", "Ölçek",
     "Solar panel grup boyutu kWp'ye göre", "Kolay", 3),
    ("12.7", "12·Çevre", "Yenilenebilir oranı (%)", "Etiket",
     "Yeşil yüzde göstergesi", "Kolay", 2),
    ("12.8", "12·Çevre", "Biyogaz potansiyeli", "Yapı",
     "Biyogaz tesisi (kubbeli yapı, alev animasyonu)", "Orta", 3),
    ("12.9", "12·Çevre", "Atık ayrıştırma", "Yapı",
     "Renkli atık konteynerleri (3-4 adet)", "Kolay", 2),
    ("12.13", "12·Çevre", "Karbon kredi sistemi", "Etiket",
     "Premium yeşil rozet (Verified)", "Kolay", 2),

    # --- Section 13: HEDEF ---
    ("13.2", "13·Hedef", "5y sürü büyüklük hedefi", "Vizyon",
     "Faded/transparent silüetlerle gelecek ahır (hayalet görünüm)", "Zor", 2),
    ("13.4", "13·Hedef", "2y yatırım planı (çoklu)", "İkon",
     "Vinç + yapım levhası ('Yapım aşamasında')", "Orta", 2),
    ("13.6", "13·Hedef", "TKDK/IPARD desteği", "Etiket",
     "AB/devlet destekli rozeti", "Kolay", 2),

    # --- Section 14: RİSK ---
    ("14.3", "14·Risk", "Yedek jeneratör", "Yapı",
     "Çiftlik yanında yeşil jeneratör ünitesi", "Orta", 3),
    ("14.4", "14·Risk", "Yangın söndürme sistemi", "Donanım",
     "Ahır üstünde kırmızı sprinkler/yangın söndürücüleri", "Orta", 2),
    ("14.5", "14·Risk", "Biyogüvenlik protokolü", "İkon",
     "Çiftlik girişinde dezenfeksiyon istasyonu", "Orta", 2),
    ("14.6", "14·Risk", "Karantina alanı", "Yapı",
     "Ana ahırdan uzakta küçük ayrı ahır (çitli)", "Orta", 3),
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Görsel Seçim"

    # Header
    headers = ["SEÇ ✓", "Soru No", "Bölüm", "Soru", "Kategori",
               "Önerilen Görsel / Animasyon", "Karmaşıklık", "Öncelik (1-5)", "Notlar"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        c.fill = PatternFill("solid", start_color="14532D")  # dark emerald
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(
            bottom=Side(style="medium", color="0F3D24"),
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD")
        )

    # Header row height
    ws.row_dimensions[1].height = 38

    # Data rows
    for r_idx, (no, sec, soru, kat, gorsel, karm, oncelik) in enumerate(ROWS, 2):
        ws.cell(r_idx, 1, "").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r_idx, 2, no).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r_idx, 3, sec).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(r_idx, 4, soru).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(r_idx, 5, kat).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r_idx, 6, gorsel).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(r_idx, 7, karm).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r_idx, 8, oncelik).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r_idx, 9, "").alignment = Alignment(vertical="center", wrap_text=True)

        # Section row color tint
        section_colors = {
            "1·": "F5F3FF", "2·": "FEF3C7", "3·": "DCFCE7", "4·": "FCE7F3",
            "5·": "DBEAFE", "6·": "ECFCCB", "7·": "EDE9FE", "8·": "FFEDD5",
            "9·": "FEF9C3", "10·": "CCFBF1", "11·": "DBEAFE", "12·": "DCFCE7",
            "13·": "FCE7F3", "14·": "F1F5F9",
        }
        prefix = sec.split("·")[0] + "·"
        bg = section_colors.get(prefix, "FFFFFF")
        for col in range(2, 10):
            ws.cell(r_idx, col).fill = PatternFill("solid", start_color=bg)
            ws.cell(r_idx, col).border = Border(
                top=Side(style="thin", color="E5E7EB"),
                bottom=Side(style="thin", color="E5E7EB"),
                left=Side(style="thin", color="E5E7EB"),
                right=Side(style="thin", color="E5E7EB"),
            )

        # SEÇ column — special bordered cell (clickable feel)
        sec_cell = ws.cell(r_idx, 1)
        sec_cell.fill = PatternFill("solid", start_color="FFFFFF")
        sec_cell.font = Font(bold=True, size=14, color="16A34A", name="Calibri")
        sec_cell.border = Border(
            top=Side(style="medium", color="86EFAC"),
            bottom=Side(style="medium", color="86EFAC"),
            left=Side(style="medium", color="86EFAC"),
            right=Side(style="medium", color="86EFAC"),
        )

        # Karmaşıklık color
        karm_cell = ws.cell(r_idx, 7)
        karm_colors = {"Kolay": "86EFAC", "Orta": "FCD34D", "Zor": "FCA5A5"}
        karm_cell.fill = PatternFill("solid", start_color=karm_colors.get(karm, "FFFFFF"))
        karm_cell.font = Font(bold=True, size=10, color="1F2937")

        # Öncelik with star symbol
        onc_cell = ws.cell(r_idx, 8)
        stars = "★" * oncelik + "☆" * (5 - oncelik)
        onc_cell.value = stars
        priority_colors = {5: "DC2626", 4: "EA580C", 3: "CA8A04", 2: "6B7280", 1: "9CA3AF"}
        onc_cell.font = Font(bold=True, size=11, color=priority_colors.get(oncelik, "000000"))

    # Column widths
    widths = {1: 8, 2: 9, 3: 14, 4: 35, 5: 12, 6: 60, 7: 13, 8: 11, 9: 24}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Row heights — auto-fit-ish
    for r in range(2, len(ROWS) + 2):
        ws.row_dimensions[r].height = 42

    # Freeze header
    ws.freeze_panes = "A2"

    # Autofilter
    ws.auto_filter.ref = f"A1:I{len(ROWS) + 1}"

    # Summary on top
    ws.cell(len(ROWS) + 3, 1, "TOPLAM").font = Font(bold=True, size=12)
    ws.cell(len(ROWS) + 3, 2, f"=COUNTA(A2:A{len(ROWS)+1})").font = Font(bold=True, size=12, color="16A34A")
    ws.cell(len(ROWS) + 3, 3, f"/ {len(ROWS)} seçildi").font = Font(size=11, color="6B7280")

    # Instructions sheet
    ws2 = wb.create_sheet("Kullanım")
    ws2.column_dimensions['A'].width = 100
    ws2['A1'] = "📋 Dijital İkiz Görsel Seçim Listesi — Kullanım"
    ws2['A1'].font = Font(bold=True, size=16, color="14532D")
    instr = [
        "",
        "1. 'Görsel Seçim' sekmesine git.",
        "2. İstediğin her satırın SEÇ ✓ kolonuna 'X' veya '✓' yaz (ya da herhangi bir karakter).",
        "3. Karmaşıklık: Kolay = 1-2 saat, Orta = yarım gün, Zor = 1+ gün iş.",
        "4. Öncelik: 5★ = mutlaka, 1★ = opsiyonel.",
        "5. Notlar kolonuna özel istek yazabilirsin (örn: 'Sadece pikselart stili').",
        "6. Filtre kullan: kategori bazlı veya öncelik bazlı seçim yapabilirsin.",
        "",
        "Toplam görselleştirilebilir soru: " + str(len(ROWS)),
        "Mevcut entegrasyon: 8 yapı (ahır, sağım, tank, silo, enerji, su, gübre, sürü).",
        "",
        "Dosyayı kaydet ve bana gönder, ben tüm seçili olanları kodlarım.",
        "",
        "KATEGORİLER:",
        "• Ana Yapı: Büyük binalar (ahır, sağım odası)",
        "• Yapı: Yardımcı binalar (silo, tank, biyogaz, depo)",
        "• Donanım: Sensör, ekipman, otomasyon",
        "• Hayvan: İnek/buzağı/boğa figürleri",
        "• Kişi: Çalışan/vet/müdür figürleri",
        "• Araç: Traktör/tanker/kamyon",
        "• Tarla: Mısır/yonca/saman alanları",
        "• Etiket: Köşede badge/rozet/sayaç",
        "• Animasyon: Hareketli efektler (duş, akış, dönen vs.)",
        "• Atmosfer: Gündüz/gece, hava efektleri",
        "• Vizyon: Hedef/gelecek silüetleri",
        "• İkon: Küçük simge/indicator",
        "• Ölçek: Boyut değişimi (kapasite/alan'a göre)",
        "• Yerleşim: Ahır içi bölme/zoning",
        "• Altyapı: Elektrik direği, wifi kulesi, internet",
        "• Konum: Harita/GPS",
        "• Detay: Renk/doku/aydınlatma farkı",
        "• Sağlık: Yeşil/sarı/kırmızı durum göstergesi",
    ]
    for i, line in enumerate(instr, 2):
        ws2.cell(i, 1, line).alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("•") or line.startswith("KATEGORİ"):
            ws2.cell(i, 1).font = Font(name="Calibri", size=11)
        if line.startswith("KATEGORİLER"):
            ws2.cell(i, 1).font = Font(bold=True, size=12, color="14532D")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"Rows: {len(ROWS)}")
    print(f"Sections covered: {len(set(r[1] for r in ROWS))}")


if __name__ == "__main__":
    main()
