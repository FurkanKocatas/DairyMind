"""Extract DairyMind Çiftlik Anketi xlsx and merge with overrides.json into questions.json."""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT.parent / "DairyMind Cifftlik Anketi.xlsx"
OVERRIDES = ROOT / "data" / "overrides.json"
OUT = ROOT / "data" / "questions.json"


SECTION_META = {
    "1.İşletme Kimliği": {"id": 1, "title": "İşletme Kimliği & Yasal Yapı", "icon": "🏢", "color": "indigo"},
    "2.Arazi Altyapı": {"id": 2, "title": "Arazi & Altyapı", "icon": "🏗️", "color": "amber"},
    "3.Sürü Yapısı": {"id": 3, "title": "Sürü Yapısı", "icon": "🐄", "color": "emerald"},
    "4.Sağlık-Üreme": {"id": 4, "title": "Hayvan Sağlığı & Üreme", "icon": "🏥", "color": "rose"},
    "5.Süt Üretimi": {"id": 5, "title": "Süt Üretimi & Kalite", "icon": "🥛", "color": "sky"},
    "6.Beslenme": {"id": 6, "title": "Beslenme & Yem Sistemi", "icon": "🌾", "color": "lime"},
    "7.Sağım Sistemi": {"id": 7, "title": "Sağım Sistemi", "icon": "🤖", "color": "violet"},
    "8.İşgücü": {"id": 8, "title": "İşgücü & Yönetim", "icon": "👥", "color": "orange"},
    "9.Finansal": {"id": 9, "title": "Finansal Yapı", "icon": "💰", "color": "yellow"},
    "10.Pazarlama": {"id": 10, "title": "Pazarlama & Satış", "icon": "📈", "color": "teal"},
    "11.Teknoloji": {"id": 11, "title": "Teknoloji & Dijitalleşme", "icon": "💻", "color": "blue"},
    "12.Sürdürülebilirlik": {"id": 12, "title": "Çevre & Sürdürülebilirlik", "icon": "♻️", "color": "green"},
    "13.Hedef Strateji": {"id": 13, "title": "Hedef & Strateji", "icon": "🎯", "color": "fuchsia"},
    "14.Risk Operasyon": {"id": 14, "title": "Risk & Operasyonel", "icon": "⚙️", "color": "slate"},
}


def detect_default_type(yanit_tipi):
    yt = (yanit_tipi or "").strip().lower()
    if yt == "sayı":
        return "number"
    if yt == "tek seçim":
        return "select_text"
    if yt == "çoklu":
        return "multi_text"
    return "text"


def make_qid(section_id, soru_no):
    return f"q_{section_id}_{soru_no.replace('.', '_')}"


def main():
    if not XLSX.exists():
        print(f"xlsx not found: {XLSX}", file=sys.stderr)
        sys.exit(1)

    overrides = {}
    if OVERRIDES.exists():
        with OVERRIDES.open(encoding="utf-8") as f:
            overrides = json.load(f)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        if sheet_name not in SECTION_META:
            continue
        ws = wb[sheet_name]
        meta = SECTION_META[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        description = ""
        if len(rows) > 1 and rows[1] and rows[1][0]:
            description = str(rows[1][0]).strip()

        questions = []
        for row in rows[4:]:
            if not row or not row[0]:
                continue
            try:
                int(row[0])
            except (TypeError, ValueError):
                continue
            soru_no = str(row[1] or "").strip()
            soru = str(row[2] or "").strip()
            yanit_tipi = str(row[3] or "").strip()
            secenekler = str(row[4] or "").strip() if row[4] else ""
            onem = str(row[6] or "").strip() if len(row) > 6 and row[6] else ""
            modul = str(row[7] or "").strip() if len(row) > 7 and row[7] else ""

            qid = make_qid(meta["id"], soru_no)

            q = {
                "id": qid,
                "no": soru_no,
                "text": soru,
                "type": detect_default_type(yanit_tipi),
                "options": [],
                "unit": "",
                "placeholder": secenekler,
                "help": "",
                "importance": onem,
                "module": modul,
                "raw_yanit_tipi": yanit_tipi,
                "raw_secenekler": secenekler,
                "validation": None,
                "depends_on": None,
                "min": None,
                "max": None,
                "step": None,
            }

            # Number default uses Seçenekler as unit hint
            if q["type"] == "number" and secenekler:
                q["unit"] = secenekler
                q["placeholder"] = ""

            # Apply overrides
            ov = overrides.get(qid)
            if ov:
                for k, v in ov.items():
                    if k.startswith("_"):
                        continue
                    q[k] = v

            questions.append(q)

        sections.append({
            "id": meta["id"],
            "slug": f"bolum-{meta['id']}",
            "title": meta["title"],
            "icon": meta["icon"],
            "color": meta["color"],
            "description": description,
            "questions": questions,
            "question_count": len(questions),
        })

    out = {
        "title": "DairyMind™ Çiftlik Dijital İkiz Anketi",
        "description": "Süt çiftliğinizi dijital ikize aktarmak için kapsamlı veri toplama anketi.",
        "version": "1.1",
        "total_sections": len(sections),
        "total_questions": sum(s["question_count"] for s in sections),
        "sections": sections,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    # Stats
    type_counts = {}
    for s in sections:
        for q in s["questions"]:
            type_counts[q["type"]] = type_counts.get(q["type"], 0) + 1
    print(f"Wrote {OUT}  sections={len(sections)}  questions={out['total_questions']}")
    print("Types:", type_counts)
    print(f"Overrides applied: {sum(1 for s in sections for q in s['questions'] if q['id'] in overrides)}/{out['total_questions']}")


if __name__ == "__main__":
    main()
